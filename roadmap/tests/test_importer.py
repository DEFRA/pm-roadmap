import io

import openpyxl
from django.test import TestCase

from roadmap.models import Roadmap, Item, Tag, TAG_COLOUR_PALETTE
from roadmap.importers import import_items, build_template_workbook

FULL_HEADER = [
    'Row ID', 'Item Type', 'Title', 'Description', 'Start Date', 'End Date',
    'Priority', 'Size', 'PRD Link', 'Backlog Link',
    'Defra Outcomes', 'Gov Objectives', 'Objectives', 'Teams', 'Categories',
    'Linked Activities',
]
# An older template that predates the Objectives column.
LEGACY_HEADER = [h for h in FULL_HEADER if h != 'Objectives']


def make_workbook(roadmap_id, header, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'Roadmap ID'
    ws['B1'] = roadmap_id
    ws.append([None])          # row 2 blank
    ws.append(header)          # row 3 headers
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


class TemplateTests(TestCase):
    def test_template_has_objectives_column(self):
        ws = build_template_workbook().active
        headers = [c.value for c in ws[3]]
        self.assertIn('Objectives', headers)


class ImportTests(TestCase):
    def setUp(self):
        self.rm = Roadmap.objects.create(name='Import RM', roadmap_type=Roadmap.SERVICE)
        # Central tags must pre-exist.
        Tag.objects.create(name='Grow Revenue', tag_type=Tag.OUTCOME)
        Tag.objects.create(name='Development', tag_type=Tag.CATEGORY)

    def _row(self, title='Activity One', objectives='', teams='', outcomes='Grow Revenue',
             cats='Development', item_type='activity'):
        return ['1', item_type, title, 'desc', '01/07/2026', '30/09/2026', 'high', 'M',
                '', '', outcomes, '', objectives, teams, cats, '']

    def test_central_tag_linked_scoped_autocreated(self):
        buf = make_workbook(self.rm.pk, FULL_HEADER,
                            [self._row(objectives='Faster Permits', teams='Marine Ops')])
        res = import_items(buf)
        self.assertEqual(res['created'], 1)
        self.assertEqual(res['errors'], [])

        item = Item.objects.get(title='Activity One')
        tag_names = {t.name for t in item.tags.all()}
        self.assertEqual(tag_names, {'Grow Revenue', 'Development', 'Faster Permits', 'Marine Ops'})

        # Scoped tags created against this roadmap, with a palette colour.
        obj = Tag.objects.get(name='Faster Permits', tag_type=Tag.OBJECTIVE)
        self.assertEqual(obj.roadmap_id, self.rm.pk)
        self.assertIn(obj.colour, TAG_COLOUR_PALETTE)

    def test_imported_tags_synced_to_roadmap_membership(self):
        buf = make_workbook(self.rm.pk, FULL_HEADER, [self._row(teams='Marine Ops')])
        import_items(buf)
        names = set(self.rm.tags.values_list('name', flat=True))
        self.assertIn('Marine Ops', names)
        self.assertIn('Grow Revenue', names)

    def test_missing_central_tag_reports_error(self):
        buf = make_workbook(self.rm.pk, FULL_HEADER, [self._row(outcomes='Does Not Exist')])
        res = import_items(buf)
        self.assertTrue(any('Does Not Exist' in e for e in res['errors']))

    def test_missing_title_skipped(self):
        buf = make_workbook(self.rm.pk, FULL_HEADER, [self._row(title='')])
        res = import_items(buf)
        self.assertEqual(res['created'], 0)
        self.assertEqual(res['skipped'], 1)

    def test_invalid_item_type_skipped(self):
        buf = make_workbook(self.rm.pk, FULL_HEADER, [self._row(item_type='bogus')])
        res = import_items(buf)
        self.assertEqual(res['skipped'], 1)

    def test_update_existing_item_by_title(self):
        Item.objects.create(roadmap=self.rm, item_type=Item.ACTIVITY, title='Activity One')
        buf = make_workbook(self.rm.pk, FULL_HEADER, [self._row()])
        res = import_items(buf)
        self.assertEqual((res['created'], res['updated']), (0, 1))

    def test_legacy_template_without_objectives_column(self):
        # Build a row matching the legacy header (no Objectives cell).
        row = ['1', 'activity', 'Legacy Item', 'd', '01/07/2026', '30/09/2026', 'high', 'M',
               '', '', 'Grow Revenue', '', 'Marine Ops', 'Development', '']
        buf = make_workbook(self.rm.pk, LEGACY_HEADER, [row])
        res = import_items(buf)
        self.assertEqual(res['created'], 1)
        self.assertEqual(res['errors'], [])

    def test_bad_roadmap_id(self):
        buf = make_workbook(999999, FULL_HEADER, [self._row()])
        res = import_items(buf)
        self.assertTrue(res['errors'])
        self.assertEqual(res['created'], 0)
