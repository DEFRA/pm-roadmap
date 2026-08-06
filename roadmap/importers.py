"""Spreadsheet importer for roadmap items."""
from __future__ import annotations

import openpyxl
from datetime import datetime, date

from .models import Roadmap, Item, Tag

# ── Column names exactly as they appear in row 3 of the template ──────────────
COL_ROW_ID       = 'Row ID'
COL_ITEM_TYPE    = 'Item Type'
COL_TITLE        = 'Title'
COL_DESCRIPTION  = 'Description'
COL_START_DATE   = 'Start Date'
COL_END_DATE     = 'End Date'
COL_PRIORITY     = 'Priority'
COL_SIZE         = 'Size'
COL_PRD_LINK     = 'PRD Link'
COL_BACKLOG_LINK = 'Backlog Link'
COL_OUTCOMES     = 'Defra Outcomes'
COL_GOV_OBJS     = 'Gov Objectives'
COL_OBJECTIVES   = 'Objectives'
COL_TEAMS        = 'Teams'
COL_CATEGORIES   = 'Categories'
COL_LINKED       = 'Linked Activities'

VALID_ITEM_TYPES = {c[0] for c in Item.ITEM_TYPE_CHOICES}
# Accept the display label "Key Result" (and legacy "metric") for the same stored value.
ITEM_TYPE_ALIASES = {'key result': Item.METRIC, 'key_result': Item.METRIC}
VALID_PRIORITIES = {c[0] for c in Item.PRIORITY_CHOICES} | {''}
VALID_SIZES      = {c[0] for c in Item.SIZE_CHOICES} | {''}


def _str(val) -> str:
    return str(val).strip() if val is not None else ''


def _parse_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, (date, datetime)):
        return val if isinstance(val, date) else val.date()
    s = _str(val)
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _parse_tags(raw: str, tag_type: str, row_num: int, errors: list, roadmap=None) -> list[Tag]:
    """Resolve comma-separated tag names to Tag objects.

    Scoped types (teams, objectives) are looked up within this roadmap and
    auto-created if missing (they are roadmap-local). Central types (outcomes,
    gov objectives, categories) must already exist — they are admin-defined.
    """
    scoped = tag_type in Tag.SCOPED_TYPES
    tags = []
    for name in (n.strip() for n in raw.split(',') if n.strip()):
        lookup = {'name': name, 'tag_type': tag_type, 'roadmap': roadmap if scoped else None}
        try:
            tags.append(Tag.objects.get(**lookup))
        except Tag.DoesNotExist:
            if scoped and roadmap is not None:
                # colour omitted → model assigns a random palette colour
                tags.append(Tag.objects.create(**lookup))
            else:
                errors.append(
                    f"Row {row_num}: tag \"{name}\" not found "
                    f"(type: {tag_type}). Create it in admin first."
                )
    return tags


def import_items(file_obj) -> dict:
    """
    Parse an .xlsx file and create/update items on the specified roadmap.
    Returns a dict with keys: roadmap, created, updated, skipped, errors, warnings.
    """
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
    except Exception as exc:
        return {'errors': [f"Could not open file: {exc}"], 'created': 0,
                'updated': 0, 'skipped': 0, 'warnings': []}

    ws = wb.active

    # ── Roadmap ID from cell B1 ───────────────────────────────────────────────
    roadmap_id = ws['B1'].value
    if not roadmap_id:
        return {'errors': ["Cell B1 must contain the Roadmap ID."],
                'created': 0, 'updated': 0, 'skipped': 0, 'warnings': []}
    try:
        roadmap = Roadmap.objects.get(pk=int(roadmap_id))
    except (Roadmap.DoesNotExist, ValueError):
        return {'errors': [f"No roadmap found with ID {roadmap_id}."],
                'created': 0, 'updated': 0, 'skipped': 0, 'warnings': []}

    # ── Auto-detect header row (scan rows 2-5 for the row containing 'Title') ──
    header_row_num = None
    for r in range(2, 6):
        vals = [_str(c.value) for c in ws[r]]
        if COL_TITLE in vals:
            header_row_num = r
            break
    if header_row_num is None:
        return {'errors': ["Could not find a header row containing 'Title' in rows 2–5. "
                           "Please use the official template."],
                'created': 0, 'updated': 0, 'skipped': 0, 'warnings': []}

    header_row = [_str(c.value) for c in ws[header_row_num]]

    # Pre-build index of column positions
    try:
        idx = {name: header_row.index(name) for name in [
            COL_ROW_ID, COL_ITEM_TYPE, COL_TITLE, COL_DESCRIPTION,
            COL_START_DATE, COL_END_DATE, COL_PRIORITY, COL_SIZE,
            COL_PRD_LINK, COL_BACKLOG_LINK,
            COL_OUTCOMES, COL_GOV_OBJS, COL_TEAMS, COL_CATEGORIES, COL_LINKED,
        ]}
    except ValueError as exc:
        return {'errors': [f"Missing expected column: {exc}. "
                           "Please use the official template."],
                'created': 0, 'updated': 0, 'skipped': 0, 'warnings': []}

    # Objectives (service objectives) is optional — older templates won't have it.
    if COL_OBJECTIVES in header_row:
        idx[COL_OBJECTIVES] = header_row.index(COL_OBJECTIVES)

    errors: list[str] = []
    warnings: list[str] = []
    created = updated = skipped = 0

    row_id_to_item: dict[str, Item] = {}
    membership_tags: set[Tag] = set()

    # ── First pass: create / update items ────────────────────────────────────
    for row_num, row in enumerate(ws.iter_rows(min_row=header_row_num + 1, values_only=True), start=4):
        if not any(row):
            continue  # blank row — skip

        def v(col_name):
            i = idx.get(col_name)
            return row[i] if i is not None else None

        row_id    = _str(v(COL_ROW_ID))
        item_type = _str(v(COL_ITEM_TYPE)).lower()
        item_type = ITEM_TYPE_ALIASES.get(item_type, item_type)
        title     = _str(v(COL_TITLE))

        # Required fields
        if not title:
            errors.append(f"Row {row_num}: Title is required — row skipped.")
            skipped += 1
            continue
        if item_type not in VALID_ITEM_TYPES:
            errors.append(
                f"Row {row_num}: \"{item_type}\" is not a valid Item Type "
                f"(use: activity, milestone, key result) — row skipped."
            )
            skipped += 1
            continue

        priority = _str(v(COL_PRIORITY)).lower()
        size     = _str(v(COL_SIZE)).upper()

        if priority and priority not in VALID_PRIORITIES:
            errors.append(f"Row {row_num}: invalid Priority \"{priority}\" — ignored.")
            priority = ''
        if size and size not in VALID_SIZES:
            errors.append(f"Row {row_num}: invalid Size \"{size}\" — ignored.")
            size = ''

        # Tags — collected but errors non-fatal (item still saved without bad tags).
        # Teams + Objectives are roadmap-scoped (auto-created); others are central.
        outcome_tags   = _parse_tags(_str(v(COL_OUTCOMES)),   Tag.OUTCOME,       row_num, errors)
        gov_tags       = _parse_tags(_str(v(COL_GOV_OBJS)),   Tag.GOV_OBJECTIVE, row_num, errors)
        objective_tags = _parse_tags(_str(v(COL_OBJECTIVES)), Tag.OBJECTIVE,     row_num, errors, roadmap=roadmap)
        team_tags      = _parse_tags(_str(v(COL_TEAMS)),      Tag.ORGANISATION,  row_num, errors, roadmap=roadmap)
        category_tags  = _parse_tags(_str(v(COL_CATEGORIES)), Tag.CATEGORY,      row_num, errors)
        all_tags       = outcome_tags + gov_tags + objective_tags + team_tags + category_tags
        membership_tags.update(all_tags)

        # Dates
        start_date = _parse_date(v(COL_START_DATE))
        end_date   = _parse_date(v(COL_END_DATE))
        if v(COL_START_DATE) and start_date is None:
            warnings.append(f"Row {row_num}: could not parse Start Date \"{v(COL_START_DATE)}\" — left blank.")
        if v(COL_END_DATE) and end_date is None:
            warnings.append(f"Row {row_num}: could not parse End Date \"{v(COL_END_DATE)}\" — left blank.")

        # Create or update (match on roadmap + title)
        item, was_created = Item.objects.update_or_create(
            roadmap=roadmap,
            title=title,
            defaults={
                'item_type':    item_type,
                'description':  _str(v(COL_DESCRIPTION)),
                'start_date':   start_date,
                'end_date':     end_date,
                'priority':     priority,
                'size':         size,
                'prd_link':     _str(v(COL_PRD_LINK)),
                'backlog_link': _str(v(COL_BACKLOG_LINK)),
            },
        )
        item.tags.set(all_tags)

        if was_created:
            created += 1
        else:
            updated += 1

        if row_id:
            row_id_to_item[row_id] = item

    # Sync imported tags into roadmap.tags so they appear in the header (which is
    # membership-based, matching the manage-tags modal).
    if membership_tags:
        roadmap.tags.add(*membership_tags)

    # ── Second pass: wire linked_activities ───────────────────────────────────
    for row_num, row in enumerate(ws.iter_rows(min_row=header_row_num + 1, values_only=True), start=4):
        if not any(row):
            continue

        row_id     = _str(row[idx[COL_ROW_ID]])
        linked_raw = _str(row[idx[COL_LINKED]])

        if not linked_raw or not row_id or row_id not in row_id_to_item:
            continue

        item = row_id_to_item[row_id]
        linked_items = []
        for ref in (r.strip() for r in linked_raw.split(',') if r.strip()):
            if ref in row_id_to_item:
                target = row_id_to_item[ref]
                if target.item_type == Item.ACTIVITY:
                    linked_items.append(target)
                else:
                    warnings.append(
                        f"Row {row_num}: Row ID \"{ref}\" is not an activity — link skipped."
                    )
            else:
                warnings.append(
                    f"Row {row_num}: Row ID \"{ref}\" not found in this upload — link skipped."
                )
        item.linked_activities.set(linked_items)

    return {
        'roadmap': roadmap,
        'created': created,
        'updated': updated,
        'skipped': skipped,
        'errors':  errors,
        'warnings': warnings,
    }


def build_template_workbook() -> openpyxl.Workbook:
    """Return an openpyxl Workbook pre-formatted as the import template."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Items'

    NAVY  = '00001C'
    GOLD  = 'E6C300'
    GREY  = 'D4DDE8'
    LIGHT = 'F0F4F8'

    # ── Row 1: Roadmap ID config ──────────────────────────────────────────────
    ws['A1'] = 'Roadmap ID'
    ws['A1'].font = Font(bold=True, color=NAVY)
    ws['A1'].fill = PatternFill('solid', fgColor=GOLD)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws['B1'] = ''   # admin fills in the ID
    ws['B1'].fill = PatternFill('solid', fgColor='FFFFFF')
    ws.row_dimensions[1].height = 22

    # ── Row 2: blank separator ────────────────────────────────────────────────
    ws.append([None])          # actually insert a blank row so headers land in row 3
    ws.row_dimensions[2].height = 8

    # ── Row 3: column headers ─────────────────────────────────────────────────
    headers = [
        COL_ROW_ID, COL_ITEM_TYPE, COL_TITLE, COL_DESCRIPTION,
        COL_START_DATE, COL_END_DATE, COL_PRIORITY, COL_SIZE,
        COL_PRD_LINK, COL_BACKLOG_LINK,
        COL_OUTCOMES, COL_GOV_OBJS, COL_OBJECTIVES, COL_TEAMS, COL_CATEGORIES, COL_LINKED,
    ]
    ws.append(headers)
    for cell in ws[3]:
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.fill      = PatternFill('solid', fgColor=NAVY)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[3].height = 30

    # ── Row 4: example activity ───────────────────────────────────────────────
    ws.append([
        '1', 'activity', 'Example Activity', 'A short description',
        '01/06/2026', '31/08/2026', 'high', 'M',
        'https://example.com/prd', '',
        'Thriving Places', 'Productive & Regenerative Land', 'Faster Permits', 'Growth Team', 'Development', '',
    ])
    # Row 5: example milestone linked to row 1
    ws.append([
        '2', 'milestone', 'Example Milestone', '',
        '30/06/2026', '30/06/2026', '', '',
        '', '',
        '', '', '', '', 'Policy', '1',
    ])
    for row in [ws[4], ws[5]]:
        for cell in row:
            cell.fill      = PatternFill('solid', fgColor=LIGHT)
            cell.alignment = Alignment(vertical='center')
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 18

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [8, 12, 32, 40, 13, 13, 10, 7, 32, 32, 26, 26, 22, 22, 20, 20]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Notes tab ─────────────────────────────────────────────────────────────
    notes = wb.create_sheet('Notes')
    notes_data = [
        ('Column', 'Required?', 'Valid values / format'),
        ('Row ID',        'Recommended', 'Any unique value per row (e.g. 1, 2, 3). Used only to wire Linked Activities.'),
        ('Item Type',     'Required',    'activity  |  milestone  |  key result'),
        ('Title',         'Required',    'Free text. If a row with this title already exists on the roadmap it will be updated.'),
        ('Description',   'Optional',    'Free text.'),
        ('Start Date',    'Optional',    'DD/MM/YYYY'),
        ('End Date',      'Optional',    'DD/MM/YYYY'),
        ('Priority',      'Optional',    'low  |  medium  |  high  |  critical   (activities only)'),
        ('Size',          'Optional',    'S  |  M  |  L  |  XL  |  XXL   (activities only)'),
        ('PRD Link',      'Optional',    'Full URL. Activities only.'),
        ('Backlog Link',  'Optional',    'Full URL. Activities only.'),
        ('Defra Outcomes','Optional',    'Comma-separated tag names. Central — must already exist (admin-defined).'),
        ('Gov Objectives','Optional',    'Comma-separated tag names. Central — must already exist (admin-defined). Group roadmaps.'),
        ('Objectives',    'Optional',    'Comma-separated tag names. Roadmap-specific — created automatically if new. Service roadmaps.'),
        ('Teams',         'Optional',    'Comma-separated tag names. Roadmap-specific — created automatically if new.'),
        ('Categories',    'Optional',    'Comma-separated tag names. Central — must already exist (admin-defined).'),
        ('Linked Activities', 'Optional','Comma-separated Row IDs of activity rows in this upload. Milestones/key results only.'),
    ]
    for r, row in enumerate(notes_data, start=1):
        for c, val in enumerate(row, start=1):
            cell = notes.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill('solid', fgColor=NAVY)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        notes.row_dimensions[r].height = 30

    notes.column_dimensions['A'].width = 20
    notes.column_dimensions['B'].width = 14
    notes.column_dimensions['C'].width = 70

    return wb
